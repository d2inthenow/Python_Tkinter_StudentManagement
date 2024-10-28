import tkinter as tk
from tkinter import ttk, messagebox, Label,Frame,Entry,Button
from tkinter import PhotoImage
import sqlite3
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from email.utils import parsedate_to_datetime
from email.header import decode_header
import email
from email.message import EmailMessage
import requests
from rasa_sdk import Action
from smtplib import SMTP_SSL
import imaplib
import pandas as pd
from datetime import datetime, timedelta
import seaborn as sns
import matplotlib.pyplot as plt
import pytz

# Main window
window = tk.Tk()
window.geometry("1350x700") 
window.title("Student Management System")
window.withdraw()

# Connect to the database
def create_table():
    try:
        connection = sqlite3.connect('student.db')
        TABLE_NAME = "STUDENT"
        SCHOOLYEAR = "dot"
        CLASSCODE = "malop"
        SUBJECT = "monhoc"
        NAME = "hoten"
        STUDENTCODE = "mssv"
        ABSENCEWITHPERMISSION="vangphep"
        ABSENCEWITHOUTPERMISSION="vangkhongphep"
        TOTALNUMBEROFLESSONS="tongsotiet"
        PERCENTABSENT ="phantramvang"
        DATE="ngayvang"
        
        #Create table if it doesn't exist, with proper field types and constraints
        connection.execute(f'''
            CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
                {SCHOOLYEAR} TEXT,
                {CLASSCODE} TEXT,
                {SUBJECT} TEXT,
                {NAME} TEXT NOT NULL,
                {STUDENTCODE} TEXT ,
                {ABSENCEWITHPERMISSION} INTEGER DEFAULT 0,
                {ABSENCEWITHOUTPERMISSION} INTEGER DEFAULT 0,
                {TOTALNUMBEROFLESSONS}  INTEGER DEFAULT 0,
                {PERCENTABSENT} REAL DEFAULT 0.0,
                {DATE} TEXT

            );
        ''')
        
        connection.commit()
        connection.close()  
        print("Table created successfully")

    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error creating table: {e}")

# Call the function to create the table
create_table()

absence_statistics = {}

def load_data_from_selected_excel():
    global absence_statistics  # Use the global variable to store absence statistics
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
    )
    if file_path:  
        try:
            connection = sqlite3.connect('student.db')
            cursor = connection.cursor()

            print(f"Trying to read: {file_path}")
            try:
                workbook = load_workbook(filename=file_path)
                sheet = workbook.active
                dot = sheet['C6'].value
                malop = sheet['C10'].value
                monhoc = sheet['C9'].value

                # Initialize absence counts
                total_excused = 0
                total_unexcused = 0
                less_than_3_absences = 0
                three_or_more_absences = 0

                for row in range(14, sheet.max_row + 1):
                    hoten = f"{sheet[f'C{row}'].value} {sheet[f'D{row}'].value}"
                    mssv = sheet[f'B{row}'].value
                    ngayvang = []
                    for col in range(4, 23):
                        value = sheet.cell(row=row, column=col).value
                        if value == 'P':  # Excused absence
                            absence_date = sheet.cell(row=12, column=col).value
                            ngayvang.append(absence_date)
                            total_excused += 1
                        elif value == 'K':  # Unexcused absence
                            absence_date = sheet.cell(row=12, column=col).value
                            ngayvang.append(absence_date)
                            total_unexcused += 1
                    
                    # Calculate total absences for this student
                    total_absences = len(ngayvang)
                    if total_absences < 3:
                        less_than_3_absences += 1
                    else:
                        three_or_more_absences += 1

                    ngayvang_str = ', '.join(map(str, ngayvang)) 
                    vangphep = sheet[f'Y{row}'].value
                    vangkhongphep = sheet[f'Z{row}'].value
                    phantramvang = sheet[f'AB{row}'].value
                    tongsotiet = sheet[f'AA{row}'].value

                    if mssv:
                        cursor.execute('''
                            INSERT OR IGNORE INTO STUDENT (dot, malop, monhoc, hoten, mssv, vangphep, vangkhongphep, tongsotiet, phantramvang, ngayvang)
                            VALUES (?, ?, ?, ?, ?,?,?,?,?,?)
                        ''', (dot, malop, monhoc, hoten, mssv, vangphep, vangkhongphep, tongsotiet, phantramvang, ngayvang_str))

                # Save the counts to the global variable
                absence_statistics = {
                    'total_excused': total_excused,
                    'total_unexcused': total_unexcused,
                    'less_than_3_absences': less_than_3_absences,
                    'three_or_more_absences': three_or_more_absences
                }

                print(f"Loaded {file_path} successfully.")
                connection.commit()
                messagebox.showinfo("Success", f"Data loaded successfully from '{file_path}'!")
                load_data() 
            except Exception as e:
                print(f"Error reading {file_path}: {e}")

            connection.close()
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Error loading data from Excel: {e}")
def show_absence():
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        
        # Giả sử bạn đang lưu mã sinh viên được chọn trong biến `selected_student_code`
        mssv = selected_student_code
        
        # Truy xuất tổng số ngày vắng và danh sách ngày vắng từ database
        cursor.execute('''
            SELECT vangphep, vangkhongphep, ngayvang 
            FROM STUDENT 
            WHERE mssv = ?
        ''', (mssv,))
        result = cursor.fetchone()  # Lấy kết quả
        
        if result:
            vangphep, vangkhongphep, ngayvang = result
            # Tính tổng số buổi vắng
            tong_so_ngay_vang = vangphep + vangkhongphep
            
            # Hiển thị kết quả cùng danh sách ngày vắng
            if ngayvang:
                ngayvang_list = ngayvang.split(", ")  # Tách chuỗi thành danh sách ngày
                ngayvang_str = "\n".join(ngayvang_list)  # Hiển thị mỗi ngày trên 1 dòng
            else:
                ngayvang_str = "Không có ngày vắng nào được lưu."
            
            messagebox.showinfo("Tổng số buổi vắng", 
                                f"Sinh viên {mssv} đã vắng tổng cộng {tong_so_ngay_vang} buổi.\n"
                                f"Các ngày vắng:\n{ngayvang_str}")
        else:
            messagebox.showwarning("Không tìm thấy", "Không tìm thấy sinh viên với mã số đã chọn.")
        
        connection.close()
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Lỗi khi truy xuất dữ liệu: {e}")

def sort_and_display_students():
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()

        # Truy vấn sắp xếp sinh viên theo tổng số buổi vắng mà không hiển thị cột tổng số buổi vắng
        cursor.execute('''
            SELECT dot, malop, monhoc, hoten, mssv,
                   (vangphep + vangkhongphep) AS tong_so_ngay_vang
            FROM STUDENT
            ORDER BY tong_so_ngay_vang DESC, hoten DESC, malop DESC, monhoc DESC
        ''')

        result = cursor.fetchall()  # Lấy toàn bộ kết quả đã sắp xếp
        
        # Xóa dữ liệu cũ trong student_table
        for row in student_table.get_children():
            student_table.delete(row)
        
        # Cập nhật dữ liệu mới vào student_table (không hiển thị tổng số buổi vắng)
        for student in result:
            dot, malop, monhoc, hoten, mssv, _ = student  # Bỏ qua tong_so_ngay_vang
            student_table.insert('', 'end', values=(dot, malop, monhoc, hoten, mssv))

        connection.close()
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Lỗi khi truy xuất dữ liệu: {e}")
  
def export_to_excel():
    try:
        # Tạo file Excel mới
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Student Attendance"

        # Tạo tiêu đề cột
        sheet.append(["Lớp", "Môn học", "Họ tên", "MSSV", "Số buổi vắng", "Ngày vắng"])

        # Lấy dữ liệu từ cơ sở dữ liệu
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        cursor.execute('''
            SELECT malop, monhoc, hoten, mssv, 
                   (vangphep + vangkhongphep) AS tong_so_buoi_vang, ngayvang
            FROM STUDENT
            WHERE (vangphep + vangkhongphep) > 3
        ''')
        students = cursor.fetchall()

        # Ghi dữ liệu vào file Excel
        for student in students:
            malop, monhoc, hoten, mssv, tong_so_buoi_vang, ngayvang = student
            sheet.append([malop, monhoc, hoten, mssv, tong_so_buoi_vang, ngayvang])

        # Lưu file Excel
        file_path = "student_attendance.xlsx"
        workbook.save(file_path)
        connection.close()

        # Gửi file Excel qua email
        send_email_with_attachment(file_path, "duongdongto@gmail.com")

        messagebox.showinfo("Success", f"File '{file_path}' đã được tạo và gửi qua email duongdongto@gmail.com thành công!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def send_email_with_attachment(file_path, recipient_email):
    sender_email = "duong0023@gmail.com"
    sender_password = "ibammoxasknmyttc"

    # Thiết lập nội dung email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = "Student Attendance Report"

    body = "Attached is the student attendance report."
    msg.attach(MIMEText(body, 'plain'))

    # Đính kèm file Excel
    attachment = open(file_path, "rb")
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={file_path}')
    msg.attach(part)
    attachment.close()

    try:
        # Kết nối đến server SMTP
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, recipient_email, text)
        server.quit()

        print(f"Email sent successfully to {recipient_email}")
        # messagebox.showinfo("Success", f"Email sent successfully to {recipient_email}")
    except Exception as e:
        print(f"Failed to send email: {e}")

def on_student_select(event):
    selected_item = student_table.selection()  
    selected_student = student_table.item(selected_item)
    global selected_student_code
    selected_student_code = selected_student['values'][4]  

def login():
    username = entry_username.get()
    password = entry_password.get()
    
    # Kết nối tới cơ sở dữ liệu
    connection = sqlite3.connect('student.db')
    cursor = connection.cursor()

    # Truy vấn để kiểm tra xem có tài khoản với username và password nhập vào không
    cursor.execute('''SELECT * FROM USERS WHERE username = ? AND password = ?''', (username, password))
    result = cursor.fetchone()

    connection.close()

    if result:
        messagebox.showinfo("Login Success", f"Welcome {username}!")
        login_window.destroy()
        window.deiconify()  # Mở cửa sổ chính nếu đăng nhập thành công
    else:
        messagebox.showerror("Login Failed", "Invalid username or password")

def logout():
    # Đóng cửa sổ chính và mở lại cửa sổ đăng nhập
    window.destroy()
    open_login_window()

def open_login_window():
    global login_window
    global entry_username
    global entry_password
    login_window = tk.Toplevel()  
    login_window.title("Login")
    login_window.geometry("925x500+300+200")
    login_window.configure(bg="#fff")
    login_window.resizable(False, False)

    img = PhotoImage(file="login.png")
    Label(login_window, image=img, bg="white").place(x=50, y=90)

    frame = Frame(login_window, width=350, height=350, bg="white")
    frame.place(x=480, y=70)

    heading = Label(frame, text="Sign In", font=("Times New Roman", 23, "bold"), bg="white", fg="#57a1F8")
    heading.place(x=100, y=5)

    # Sign In
    def on_enter(e):
        entry_username.delete(0,'end')

    def on_leave(e):
        if entry_username.get() == "":
            entry_username.insert(0, "Username")

    entry_username = Entry(frame, font=("Times New Roman", 11), width=25,border=0,bg='white',fg='black')
    entry_username.place(x=30, y=80)
    entry_username.insert(0, "Username")
    entry_username.bind("<FocusIn>", on_enter)
    entry_username.bind("<FocusOut>", on_leave)

    Frame(frame, width=295, height=2, bg="black").place(x=25, y=107)

    def on_enter(e):
        entry_password.delete(0,'end')

    def on_leave(e):
        if entry_password.get() == "":
            entry_password.insert(0, "Password")

    entry_password = Entry(frame, font=("Times New Roman", 11), width=25,border=0,bg='white',fg='black' ,show="*")
    entry_password.place(x=30, y=150)
    entry_password.insert(0, "Password")
    entry_password.bind("<FocusIn>", on_enter)
    entry_password.bind("<FocusOut>", on_leave)

    Frame(frame, width=295, height=2, bg="black").place(x=25, y=177)

    Button(frame, width=39, pady=7, text='Sign In', bg='#57a1f8', fg='white', border=0, command=login).place(x=35, y=204)
    
    label = Label(frame, text="Don't have an account?", font=("Times New Roman", 10), bg="white", fg="black")
    label.place(x=75, y=270)

    sign_up = Button(frame, width=6, cursor='hand2', text='Sign Up', bg='white', fg='#57a1f8', border=0, command=register)
    sign_up.place(x=215, y=270)


def register():
    register_window = tk.Toplevel()
    register_window.title("Register")
    register_window.geometry("925x500+300+200")
    register_window.configure(bg="#fff")
    register_window.resizable(False, False)
    img= PhotoImage(file="register.png")
    Label(register_window, image=img, border=0, bg="white").place(x=50, y=90)

    frame = Frame(register_window, width=350, height=350, bg="#fff")
    frame.place(x=480, y=50)

    heading = Label(frame, text="Sign Up", font=("Times New Roman", 23, "bold"), bg="white", fg="#57a1F8")
    heading.place(x=100, y=5)

    # Username
    def on_enter(e):
        entry_new_username.delete(0, 'end')

    def on_leave(e):
        if entry_new_username.get() == "":
            entry_new_username.insert(0, "Username")

    entry_new_username = tk.Entry(frame, width=25, border=0, bg='white', fg='black', font=("Times New Roman", 11))
    entry_new_username.place(x=30, y=80)
    entry_new_username.insert(0, "Username")
    entry_new_username.bind("<FocusIn>", on_enter)
    entry_new_username.bind("<FocusOut>", on_leave)

    Frame(frame, width=295, height=2, bg="black").place(x=25, y=107)

    # Password
    def on_enter(e):
        entry_new_password.delete(0, 'end')

    def on_leave(e):
        if entry_new_password.get() == "":
            entry_new_password.insert(0, "Password")

    entry_new_password = tk.Entry(frame, width=25, border=0, bg='white', fg='black', font=("Times New Roman", 11))
    entry_new_password.place(x=30, y=150)
    entry_new_password.insert(0, "Password")
    entry_new_password.bind("<FocusIn>", on_enter)
    entry_new_password.bind("<FocusOut>", on_leave)
    Frame(frame, width=295, height=2, bg="black").place(x=25, y=177)

    # Confirm Password
    def on_enter(e):
        conform_new_password.delete(0, 'end')

    def on_leave(e):
        if conform_new_password.get() == "":
            conform_new_password.insert(0, "ConformPassword")

    conform_new_password = tk.Entry(frame, width=25, border=0, bg='white', fg='black', font=("Times New Roman", 11), show="*")
    conform_new_password.place(x=30, y=220)
    conform_new_password.insert(0, "ConformPassword")
    conform_new_password.bind("<FocusIn>", on_enter)
    conform_new_password.bind("<FocusOut>", on_leave)
    Frame(frame, width=295, height=2, bg="black").place(x=25, y=247)

    # Register Button
    btn_register_user = tk.Button(frame, width=39, pady=7, text="Register", fg='white', bg='#57a1f8', 
                                  command=lambda: register_user(entry_new_username.get(), entry_new_password.get(), register_window))
    btn_register_user.place(x=35, y=280)

    # I have an account label
    label = Label(frame, text="I have an account", font=("Times New Roman", 9), bg="white", fg="black")
    label.place(x=90, y=320)  # Điều chỉnh x và y để căn chỉnh chữ

    # Sign in Button
    sign_in = Button(frame, width=6, cursor='hand2', text='Sign In', bg='white', fg='#57a1f8', border=0, 
                     command=lambda: [register_window.destroy(), login()])
    sign_in.place(x=200, y=320)  # Điều chỉnh x và y để căn chỉnh nút Sign In

def register_user(new_username, new_password, register_window):
    if not new_username or not new_password:
        messagebox.showerror("Input Error", "Please fill out both username and password")
        return

    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS USERS (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL
        )''')

        cursor.execute('''INSERT INTO USERS (username, password) VALUES (?, ?)''', (new_username, new_password))
        
        connection.commit()
        connection.close()
        messagebox.showinfo("Registration Success", "User registered successfully!")
        register_window.destroy()

    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "Username already exists!")
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error registering user: {e}")

# open_login_window()
def add_student(schoolyear, classcode, subject, name, studentcode):
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        
        cursor.execute(f'''
            INSERT INTO STUDENT (dot, malop, monhoc, hoten, mssv)
            VALUES (?, ?, ?, ?, ?)
        ''', (schoolyear, classcode, subject, name, studentcode))
        

         # Gọi hàm load_data để cập nhật Treeview
        
        connection.commit()
        load_data()
        connection.close()
        
        messagebox.showinfo("Success", "Student added successfully!")
        
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "Student with this Student Code already exists!")
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error adding student: {e}")

def delete_student(studentcode):
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        
        cursor.execute(f'''
            DELETE FROM STUDENT WHERE mssv = ?
        ''', (studentcode,))
        
        if cursor.rowcount == 0:
            messagebox.showerror("Error", "Student not found!")
        else:
            messagebox.showinfo("Success", "Student deleted successfully!")
        
        connection.commit()
        load_data()
        connection.close()

    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error deleting student: {e}")
 
def update_student(schoolyear, classcode, subject, name, studentcode):
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        
        cursor.execute(f'''
            UPDATE STUDENT
            SET dot = ?,malop = ?, monhoc = ?, hoten = ?
            WHERE mssv = ?
        ''', (schoolyear, classcode, subject, name, studentcode))
        
        if cursor.rowcount == 0:
            messagebox.showerror("Error", "Student not found!")
        else:
            messagebox.showinfo("Success", "Student updated successfully!")
        
        connection.commit()
        load_data()
        connection.close()

    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error updating student: {e}")

def clear_database(): 
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()
        
        # Xóa tất cả dữ liệu trong bảng STUDENT
        cursor.execute('DELETE FROM STUDENT')
        
        connection.commit()
        load_data()
        connection.close()
        
        messagebox.showinfo("Success", "All student data cleared!")
    
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error clearing student data: {e}")

def clear_students():
   
    schoolyear.set("")
    classcode.set("")
    subject.set("")
    name.set("")
    studentcode.set("")
    messagebox.showinfo("Success", "All fields cleared!")

# hàm hiện thị dữ liệu
def load_data():
    try:
     connection = sqlite3.connect('student.db')
     cursor = connection.cursor()
     cursor.execute("SELECT * FROM STUDENT")
     rows = cursor.fetchall()
    
     # Xóa dữ liệu hiện tại trong Treeview
     student_table.delete(*student_table.get_children())
    
     # Thêm dữ liệu mới vào Treeview
     for row in rows:
        student_table.insert('', 'end', values=row)
     connection.commit()
     connection.close()
    except sqlite3.Error as e:
        print(f"Database Error: {e}")

#  hàm tìm kiếm

def search_student():
    search_term = search_entry.get()
    search_column = search_in.get()
    
    # Kiểm tra kiểu dữ liệu trước khi tìm kiếm
    if search_column == "Name":
        if any(char.isdigit() for char in search_term):  # Kiểm tra nếu có số trong tên
            messagebox.showerror("Invalid Input", "Name cannot contain numbers!")
            return
    elif search_column == "StudentCode":
        if not search_term.isalnum():  # Kiểm tra nếu mã sinh viên không phải là chữ hoặc số
            messagebox.showerror("Invalid Input", "Student Code must be alphanumeric!")
            return

    connection = sqlite3.connect('student.db')
    cursor = connection.cursor()
    
    if search_column == "Name":
        cursor.execute("SELECT * FROM STUDENT WHERE hoten LIKE ?", ('%' + search_term + '%',))
    elif search_column == "StudentCode":
        cursor.execute("SELECT * FROM STUDENT WHERE mssv LIKE ?", ('%' + search_term + '%',))
    
    rows = cursor.fetchall()
    
    student_table.delete(*student_table.get_children())
    
    if not rows:  # Kiểm tra nếu không có kết quả
        messagebox.showinfo("Search Result", "No student found!")
        load_data()  # Tải lại tất cả dữ liệu vào Treeview
    else:
        for row in rows:
            student_table.insert('', 'end', values=row)
    
    connection.commit()
    connection.close()

def send_email(recipient_list, subject, body):
    sender_email = "duong0023@gmail.com"
    sender_password = "ibammoxasknmyttc"       
    # Thiết lập nội dung email
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = ", ".join(recipient_list)

    try:
        # Kết nối đến server SMTP của nhà cung cấp email
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_list, msg.as_string())

        print(f"Email đã được gửi thành công tới: {', '.join(recipient_list)}")
        messagebox.showinfo("Success", f"Email đã được gửi thành công tới: {', '.join(recipient_list)}")

    except Exception as e:
        print(f"Lỗi khi gửi email: {e}")

def check_attendance_and_send_email():
    try:
        connection = sqlite3.connect('student.db')
        cursor = connection.cursor()

        # Truy vấn sinh viên có phantramvang >= 20%
        cursor.execute('''
            SELECT hoten, mssv, phantramvang
            FROM STUDENT
            WHERE phantramvang >= 20
        ''')
        students = cursor.fetchall()

        # Lặp qua danh sách sinh viên và kiểm tra tỷ lệ vắng
        for student in students:
            hoten, mssv, phantramvang = student
            try:
                # Kiểm tra xem phantramvang có phải là số không
                if isinstance(phantramvang, str):
                    phantramvang = phantramvang.replace(',', '.').strip()  # Thay dấu phẩy nếu có
                phantramvang = float(phantramvang)  # Chuyển đổi sang float
            except ValueError:
                print(f"Lỗi: phantramvang không hợp lệ cho sinh viên {hoten} (MSSV: {mssv})")
                continue  # Bỏ qua sinh viên này nếu không thể chuyển đổi

            # Nếu sinh viên vắng >= 50%, gửi email tới nabee2412@gmail.com
            if phantramvang >= 50:
                send_email(
                    recipient_list=["nabee2412@gmail.com"],
                    subject="Cảnh báo vắng học nghiêm trọng",
                    body=f"Sinh viên {hoten} (MSSV: {mssv}) đã vắng học {phantramvang}% thời lượng."
                )
            # Nếu sinh viên vắng >= 20% nhưng < 50%, gửi email tới duong0023@gmail.com
            elif phantramvang >= 20:
                send_email(
                    recipient_list=["duongdongto@gmail.com"],
                    subject="Cảnh báo vắng học",
                    body=f"Sinh viên {hoten} (MSSV: {mssv}) đã vắng học {phantramvang}% thời lượng."
                )

        connection.close()
    except sqlite3.Error as e:
        print(f"Lỗi truy xuất cơ sở dữ liệu: {e}")

mails_info = []

def check_reports():
    # Thông tin tài khoản email
    user_email = "duong0023@gmail.com"
    app_password = "ibammoxasknmyttc" 
    mail_server = "imap.gmail.com"
    try:
        # Kết nối tới hộp thư email bằng giao thức IMAP
        email_conn = imaplib.IMAP4_SSL(mail_server)
        email_conn.login(user_email, app_password)

        # Truy cập hộp thư "inbox"
        email_conn.select("inbox")

        # Lấy ngày bắt đầu của tháng hiện tại
        today = datetime.now()
        month_start = today.replace(day=1)
        formatted_date = month_start.strftime("%d-%b-%Y")

        # Tìm các email từ ngày đầu tháng chưa đọc và có tiêu đề "Report"
        search_criteria = f'(SINCE {formatted_date} UNSEEN)'
        result, email_ids = email_conn.search(None, search_criteria)

        if result == "OK" and email_ids[0]:
            for email_id in email_ids[0].split():
                status, email_data = email_conn.fetch(email_id, "(RFC822)")
                for part in email_data:
                    if isinstance(part, tuple):
                        message = email.message_from_bytes(part[1])

                        # Lấy người gửi và tiêu đề
                        sender_info = message.get("From")
                        title, charset = decode_header(message["Subject"])[0]
                        if isinstance(title, bytes):
                            title = title.decode(charset if charset else "utf-8")

                        # Kiểm tra tiêu đề có phải là "Report" hay không
                        if "report" in title.lower():
                            # Trích xuất thông tin từ nội dung email
                            student_email, student_name, student_id = extract_sender_data(message)
                            if student_name and student_id:
                                mails_info.append({
                                    "Email": student_email,
                                    "Tên sinh viên": student_name,
                                    "Mã sinh viên": student_id
                                })
            # Hiển thị kết quả
            show_report_results()
        else:
            messagebox.showinfo("Thông báo", "Không có email nào với tiêu đề 'report' chưa đọc trong tháng này.")

    except Exception as e:
        messagebox.showerror("Lỗi", f"Kết nối tới mail không thành công: {str(e)}")

def extract_sender_data(message):
    # Lấy nội dung email và phân tích
    if message.is_multipart():
        for content_part in message.walk():
            if content_part.get_content_type() == "text/plain":
                content_text = content_part.get_payload(decode=True).decode("utf-8")
                return analyze_email_content(content_text)
    else:
        plain_content = message.get_payload(decode=True).decode("utf-8")
        return analyze_email_content(plain_content)

def analyze_email_content(email_content):
    # Giả sử email có định dạng "Tên sinh viên_MSSV"
    try:
        content_parts = email_content.strip().split("_")
        student_id = content_parts[0]
        student_name = content_parts[1]
        sender_email = "nabee2412@gmail.com" 
        return sender_email, student_name, student_id
    except:
        return None, None, None

def show_report_results():
    # Hiển thị số lượng email có tiêu đề 'report' và tạo nút Summary
    report_count = len(mails_info)
    if report_count > 0:
        messagebox.showinfo("Thông báo", f"Đã nhận được {report_count} email có tiêu đề 'report'.")
        summary_button = tk.Button(check_frame, text="Send Summary", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=16, command=create_summary_report)
        summary_button.grid(row=0, column=5, padx=2, pady=2)
    else:
        messagebox.showinfo("Thông báo", "Không có email nào phù hợp.")

def create_summary_report():
    # Tạo file Excel chứa thông tin các email
    report_df = pd.DataFrame(mails_info)
    output_file = "report_summary.xlsx"
    report_df.to_excel(output_file, index=False)

    # Gửi file Excel qua email
    send_excel_file(output_file)

def send_excel_file(file_path):
    teacher_email = "duong0023@gmail.com"
    recipient_email = "duongdongto@gmail.com"
    email_subject = "Tổng hợp báo cáo"

    # Mật khẩu ứng dụng của email gửi
    app_password = "ibammoxasknmyttc"

    # Tạo email kèm file đính kèm
    msg = MIMEMultipart()
    msg["From"] = teacher_email
    msg["To"] = recipient_email
    msg["Subject"] = email_subject

    # Thêm file đính kèm vào email
    with open(file_path, "rb") as attachment:
        file_part = MIMEBase("application", "octet-stream")
        file_part.set_payload(attachment.read())
        encoders.encode_base64(file_part)
        file_part.add_header("Content-Disposition", f'attachment; filename="{file_path}"')
        msg.attach(file_part)

    # Kết nối tới SMTP server và gửi email
    try:
        smtp_server = SMTP_SSL("smtp.gmail.com", 465)
        smtp_server.login(teacher_email, app_password)
        smtp_server.sendmail(teacher_email, recipient_email, msg.as_string())
        smtp_server.quit()
        messagebox.showinfo("Thông báo", "File báo cáo đã được gửi thành công đến email duongdongto@gmail.com !")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Gửi email thất bại: {str(e)}")

def check_questions():
    # Thông tin đăng nhập
    email_user = 'duong0023@gmail.com'
    email_password = 'ibammoxasknmyttc'
    
    # Kết nối đến máy chủ email
    mail = imaplib.IMAP4_SSL('imap.gmail.com')
    mail.login(email_user, email_password)
    mail.select('inbox')

    # Lấy ngày đầu tháng hiện tại
    today = datetime.now()
    first_day_of_month = today.replace(day=1).strftime("%d-%b-%Y")  # Định dạng ngày phù hợp với Gmail

    # Tìm tất cả email chưa đọc trong tháng hiện tại
    search_criteria = f'(UNSEEN SINCE {first_day_of_month})'  # Tìm email chưa đọc kể từ đầu tháng
    result, data = mail.search(None, search_criteria)
    if result != 'OK':
        print("Không thể tìm kiếm email:", result)
        return

    email_ids = data[0].split()
    unanswered_questions = []
    
    for email_id in email_ids:
        result, msg_data = mail.fetch(email_id, '(RFC822)')
        msg = email.message_from_bytes(msg_data[0][1])
        
        subject = msg['Subject']
        date_str = msg['Date']
        email_date = email.utils.parsedate_to_datetime(date_str).replace(tzinfo=pytz.UTC)  # Chuyển đổi về UTC
        
        # Chuyển đổi datetime hiện tại về UTC
        current_time = datetime.now(pytz.UTC)

        if subject and 'question' in subject:
            if current_time - email_date <= timedelta(hours=24):  # So sánh datetime
                unanswered_questions.append(subject)

    question_count = len(unanswered_questions)
    messagebox.showinfo("Tổng số câu hỏi", f"Tổng số câu hỏi chưa đọc trong tháng hiện tại: {question_count}")

    if unanswered_questions:
        forward_email(unanswered_questions)

    mail.logout()

def forward_email(questions):
    to_email = 'duongdongto@gmail.com'
    subject = 'Câu hỏi chưa trả lời trong 24 giờ'
    body = 'Các câu hỏi của sinh viên trong 24 giờ qua:\n\n' + '\n'.join(questions)
    
    msg = MIMEText(body)
    msg['From'] = 'your_email@gmail.com'
    msg['To'] = to_email
    msg['Subject'] = subject
    
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login('your_email@gmail.com', 'your_password')
        server.send_message(msg)

def show_chart():
    global absence_statistics  # Access the global variable
    total_excused = absence_statistics.get('total_excused', 0)
    total_unexcused = absence_statistics.get('total_unexcused', 0)
    less_than_3_absences = absence_statistics.get('less_than_3_absences', 0)
    three_or_more_absences = absence_statistics.get('three_or_more_absences', 0)

    if total_excused == 0 and total_unexcused == 0 and less_than_3_absences == 0 and three_or_more_absences == 0:
        messagebox.showinfo("Thông báo", "Không có dữ liệu để hiển thị biểu đồ.")
        return

    # Create a DataFrame for plotting
    plot_data = pd.DataFrame({
        'Loại vắng': ['Vắng phép', 'Vắng không phép', 'Vắng ít (<3 buổi)', 'Vắng nhiều (>=3 buổi)'],
        'Số lượng': [total_excused, total_unexcused, less_than_3_absences, three_or_more_absences]
    })

    # Create a Seaborn barplot
    plt.figure(figsize=(10, 6))
    ax = sns.barplot(x='Loại vắng', y='Số lượng', data=plot_data, palette='pastel')
    plt.title('Thống kê tỷ lệ vắng có phép, vắng không phép, ít và nhiều buổi')
    plt.xlabel('Loại vắng')
    plt.ylabel('Số lượng')

    # Add data labels on top of each bar
    for p in ax.patches:
        ax.annotate(f'{p.get_height()}', (p.get_x() + p.get_width() / 2., p.get_height()),
                    ha='center', va='bottom', fontsize=12, color='black', fontweight='bold')
    plt.show()
# Login window
login_window = tk.Toplevel()
login_window.title("Login")
login_window.geometry("925x500+300+200")
login_window.configure(bg="#fff")
login_window.resizable(False, False)

img= PhotoImage(file="login.png")
Label(login_window, image=img, bg="white").place(x=50, y=90)

frame=Frame(login_window, width=350, height=350, bg="white")
frame.place(x=480, y=70)

heading=Label(frame, text="Sign In", font=("Times New Roman", 23, "bold"), bg="white", fg="#57a1F8")
heading.place(x=100, y=5)

# Sign In
def on_enter(e):
    entry_username.delete(0,'end')
def on_leave(e):
    if entry_username.get() == "":
        entry_username.insert(0, "Username")

entry_username = Entry(frame, font=("Times New Roman", 11), width=25,border=0,bg='white',fg='black')
entry_username.place(x=30, y=80)
entry_username.insert(0, "Username")
entry_username.bind("<FocusIn>", on_enter)
entry_username.bind("<FocusOut>", on_leave)

Frame(frame, width=295, height=2, bg="black").place(x=25, y=107)

def on_enter(e):
    entry_password.delete(0,'end')
def on_leave(e):
    if entry_password.get() == "":
        entry_password.insert(0, "Password")
entry_password = Entry(frame, font=("Times New Roman", 11), width=25,border=0,bg='white',fg='black' ,show="*")
entry_password.place(x=30, y=150)
entry_password.insert(0, "Password")
entry_password.bind("<FocusIn>", on_enter)
entry_password.bind("<FocusOut>", on_leave)

Frame(frame, width=295, height=2, bg="black").place(x=25, y=177)

Button(frame,width=39,pady=7,text='Sign In',bg='#57a1f8',fg='white',border=0 ,command=login).place(x=35, y=204)
label=Label(frame,text="Don't have an account?", font=("Times New Roman", 10), bg="white", fg="black")
label.place(x=75,y=270)

sign_up=Button(frame,width=6,cursor='hand2',text='Sign Up',bg='white',fg='#57a1f8',border=0,command=register)
sign_up.place(x=215, y=270)

# login_window.mainloop()

# Value
schoolyear = tk.StringVar()
classcode = tk.StringVar()
subject = tk.StringVar()
name = tk.StringVar()
studentcode = tk.StringVar()

# label and Entry
title_label = tk.Label(window, text="Student Management System", font=("Times New Roman", 26, "bold"),height=50, border=12, bg="black", foreground="red", relief=tk.GROOVE)
title_label.pack(side=tk.TOP, fill=tk.X)

detail_frame = tk.LabelFrame(window, text="Student Details", font=("Times New Roman", 20), bg="lightgrey", bd=12, relief=tk.GROOVE)
detail_frame.place(x=20, y=90, width=420, height=575)

data_frame = tk.LabelFrame(window, text="Student Data", font=("Times New Roman", 20), bg="lightgrey", bd=12, relief=tk.GROOVE)
data_frame.place(x=475, y=90, width=810, height=575)

schoolYear_label = tk.Label(detail_frame, text="Đợt", font=("Times New Roman", 15),width=15, bg="lightgrey")
schoolYear_label.grid(row=0, column=0, padx=2, pady=2)
schoolYear_ent = tk.Entry(detail_frame, bd=7, font=("Times New Roman", 15),textvariable=schoolyear)
schoolYear_ent.grid(row=0, column=1, padx=2, pady=2)

classCode_label = tk.Label(detail_frame, text="Mã lớp", font=("Times New Roman", 15),width=15, bg="lightgrey")
classCode_label.grid(row=1, column=0, padx=2, pady=2)
classCode_ent = tk.Entry(detail_frame, bd=7, font=("Times New Roman", 15),textvariable=classcode)
classCode_ent.grid(row=1, column=1, padx=2, pady=2)

subject_label = tk.Label(detail_frame, text="Tên môn học", font=("Times New Roman", 15),width=15, bg="lightgrey")
subject_label.grid(row=2, column=0, padx=2, pady=2)
subject_ent = tk.Entry(detail_frame, bd=7, font=("Times New Roman", 15),textvariable=subject)
subject_ent.grid(row=2, column=1, padx=2, pady=2)

name_label = tk.Label(detail_frame, text="Họ tên", font=("Times New Roman", 15),width=15, bg="lightgrey")
name_label.grid(row=3, column=0, padx=2, pady=2)
name_ent = tk.Entry(detail_frame, bd=7, font=("Times New Roman", 15),textvariable=name)
name_ent.grid(row=3, column=1, padx=2, pady=2)

studentCode_label = tk.Label(detail_frame, text="Mã số sinh viên", font=("Times New Roman", 15),width=15, bg="lightgrey")
studentCode_label.grid(row=4, column=0, padx=2, pady=2)
studentCode_ent = tk.Entry(detail_frame, bd=7, font=("Times New Roman", 15),textvariable=studentcode)
studentCode_ent.grid(row=4, column=1, padx=2, pady=2)

# Button (CRUD)

btn_frame = tk.Frame(detail_frame, bd=10, relief=tk.GROOVE)
btn_frame.place(x=18, y=390, width=356,height=120)

add_btn = tk.Button(btn_frame, text="Add", bd=7, font=("Times New Roman", 11), bg="lightgrey", width=11,command=lambda: add_student(schoolyear.get(), classcode.get(), subject.get(), name.get(), studentcode.get()))
add_btn.grid(row=0, column=0, padx=2, pady=2)

update_btn = tk.Button(btn_frame, text="Update", bd=7, font=("Times New Roman", 11), bg="lightgrey", width=11,command=lambda: update_student(schoolyear.get(), classcode.get(), subject.get(), name.get(), studentcode.get()))
update_btn.grid(row=0, column=1, padx=3, pady=2)

delete_btn = tk.Button(btn_frame, text="Delete", bd=7, font=("Times New Roman", 11), bg="lightgrey", width=11,command=lambda: delete_student(studentcode.get()))
delete_btn.grid(row=1, column=0, padx=2, pady=2)

clear_btn = tk.Button(btn_frame, text="Clear", bd=7, font=("Times New Roman", 11), bg="lightgrey", width=11,command=clear_students)
clear_btn.grid(row=1, column=1, padx=3, pady=2)

load_btn = tk.Button(btn_frame, text="Load Excel", bd=7, font=("Times New Roman", 11), bg="lightgrey", width=11, command=load_data_from_selected_excel)
load_btn.grid(row=0, column=2, columnspan=2, padx=2, pady=2)

clear_database_btn = tk.Button(btn_frame, text="Clear Data", bd=7, font=("Times New Roman", 11), bg="lightgrey", width=11,command=clear_database)
clear_database_btn.grid(row=1, column=2, columnspan=2, padx=2, pady=2)

# Search 

search_frame = tk.Frame(data_frame, bg="lightgrey", bd=10, relief=tk.GROOVE)
search_frame.pack(side=tk.TOP, fill=tk.X)

search_label = tk.Label(search_frame, text="Search ", font=("Times New Roman", 14), bg="lightgrey")
search_label.grid(row=0, column=0, padx=2, pady=2)

search_entry = tk.Entry(search_frame, font=("Times New Roman", 14))
search_entry.grid(row=0, column=1, padx=12, pady=2) 

search_in = ttk.Combobox(search_frame, font=("Times New Roman", 14), state="readonly", width=12)
search_in['values'] = ("Name", "StudentCode")
search_in.set("Name")
search_in.grid(row=0, column=2, padx=12, pady=2)

search_btn = tk.Button(search_frame, text="Search", font=("Times New Roman", 13), bd=9, width=33, bg="lightgrey",command=search_student)
search_btn.grid(row=0, column=3, padx=12, pady=2)


# check Day-off
check_frame = tk.Frame(data_frame, bg="lightgrey", bd=10, relief=tk.GROOVE)
check_frame.pack(side=tk.TOP, fill=tk.X)

send_email_btn = tk.Button(check_frame, text="Send Email", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=15,command=check_attendance_and_send_email)
send_email_btn.grid(row=0, column=2, padx=2, pady=2)

export_excel_btn = tk.Button(check_frame, text="Export Excel", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=15,command=export_to_excel)
export_excel_btn.grid(row=0, column=3, padx=2, pady=2)

check_button = tk.Button(check_frame, text="Check Report", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=15, command=check_reports)
check_button.grid(row=0, column=4, padx=2, pady=2)

# show data
show_frame = tk.Frame(data_frame, bg="lightgrey", bd=10, relief=tk.GROOVE)
show_frame.pack(side=tk.TOP, fill=tk.X)

day_off_btn = tk.Button(show_frame, text="Day-off", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=15,command=show_absence) 
day_off_btn.grid(row=0, column=0, padx=2, pady=2)

sort_btn = tk.Button(show_frame, text="Sort", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=15,command=sort_and_display_students)
sort_btn.grid(row=0, column=1, padx=2, pady=2)

see_chart_btn = tk.Button(show_frame, text="See Chart", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=16, command=show_chart)
see_chart_btn.grid(row=0, column=7, padx=2, pady=2)

check_question_button = tk.Button(show_frame, text="Check Question", bd=7, font=("Times New Roman", 15), bg="lightgrey", width=15, command=check_questions)
check_question_button.grid(row=0, column=6, padx=2, pady=2)



# Logout 
logout_button = tk.Button(title_label, text="Logout", bg='blue',fg='yellow',cursor='hand2', command=lambda: logout())
logout_button.pack(side=tk.RIGHT)
# Database 
main_frame = tk.Frame(data_frame,bg="lightgrey", bd=11, relief=tk.GROOVE)
main_frame.pack(fill=tk.BOTH, expand=True)

x_scroll = tk.Scrollbar(data_frame, orient=tk.HORIZONTAL)
y_scroll = tk.Scrollbar(data_frame, orient=tk.VERTICAL)

student_table = ttk.Treeview(data_frame, column=("SchoolYear", "ClassCode", "Subject", "Name", "StudentCode"), xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)

x_scroll.config(command=student_table.xview)
y_scroll.config(command=student_table.yview)

x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
y_scroll.pack(side=tk.RIGHT, fill=tk.Y)

student_table.heading("SchoolYear", text="Đợt")
student_table.heading("ClassCode", text="Mã lớp")
student_table.heading("Subject", text="Tên môn học")
student_table.heading("Name", text="Họ tên")
student_table.heading("StudentCode", text="Mã số sinh viên")

student_table['show'] = 'headings'

student_table.column("SchoolYear", width=80)
student_table.column("ClassCode", width=40)
student_table.column("Subject", width=190)
student_table.column("Name", width=110)
student_table.column("StudentCode", width=80)

student_table.pack(fill=tk.BOTH, expand=True)

def select_item(event):
    selected_item = student_table.selection()
    
    if selected_item:
        item_values = student_table.item(selected_item, 'values')
        if item_values:
         schoolyear.set(item_values[0])
         classcode.set(item_values[1])
         subject.set(item_values[2])
         name.set(item_values[3])
         studentcode.set(item_values[4])

def connect_rasa():
    url = "http://localhost:5005/webhooks/rest/webhook"
    message = {"sender": "user", "message": "Hello, bot!"}
    try:
        response = requests.post(url, json=message)
        response.raise_for_status()  # Kiểm tra nếu có lỗi HTTP
        print(response.json())
    except requests.exceptions.RequestException as e:
        print(f"Error connecting to Rasa: {e}")

class ActionReadExcel(Action):
    def name(self) -> str:
        return "action_read_excel"

    def run(self, dispatcher, tracker, domain):
        # Lấy đường dẫn đến các file Excel từ slot
        file_paths = tracker.get_slot('excel_file_paths')  # Giả sử bạn đã cung cấp slot này

        if file_paths:
            for file_path in file_paths.split(','):  # Giả sử file_paths là một chuỗi các đường dẫn cách nhau bởi dấu phẩy
                file_path = file_path.strip()  # Xóa khoảng trắng
                try:
                    connection = sqlite3.connect('student.db')
                    cursor = connection.cursor()

                    print(f"Trying to read: {file_path}")
                    try:
                        workbook = load_workbook(filename=file_path)
                        sheet = workbook.active

                        dot = sheet['C6'].value
                        malop = sheet['C10'].value
                        monhoc = sheet['C9'].value

                        # Initialize absence counts
                        total_excused = 0
                        total_unexcused = 0

                        for row in range(14, sheet.max_row + 1):
                            hoten = f"{sheet[f'C{row}'].value} {sheet[f'D{row}'].value}"
                            mssv = sheet[f'B{row}'].value
                            ngayvang = []

                            for col in range(4, 23):
                                value = sheet.cell(row=row, column=col).value
                                if value == 'P':  # Vắng có phép
                                    absence_date = sheet.cell(row=12, column=col).value
                                    ngayvang.append(absence_date)
                                    total_excused += 1
                                elif value == 'K':  # Vắng không phép
                                    absence_date = sheet.cell(row=12, column=col).value
                                    ngayvang.append(absence_date)
                                    total_unexcused += 1

                            ngayvang_str = ', '.join(map(str, ngayvang)) 
                            vangphep = sheet[f'Y{row}'].value
                            vangkhongphep = sheet[f'Z{row}'].value
                            phantramvang = sheet[f'AB{row}'].value
                            tongsotiet = sheet[f'AA{row}'].value

                            if mssv:
                                cursor.execute(''' 
                                    INSERT OR IGNORE INTO STUDENT (dot, malop, monhoc, hoten, mssv, vangphep, vangkhongphep, tongsotiet, phantramvang, ngayvang)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (dot, malop, monhoc, hoten, mssv, vangphep, vangkhongphep, tongsotiet, phantramvang, ngayvang_str))

                        # Commit the changes
                        connection.commit()

                        # Thông báo thành công
                        dispatcher.utter_message(text=f"Dữ liệu đã được tải thành công từ file {file_path}!")
                    except Exception as e:
                        print(f"Error reading {file_path}: {e}")
                        dispatcher.utter_message(text=f"Có lỗi xảy ra khi đọc file {file_path}: {str(e)}")

                    connection.close()
                except sqlite3.Error as e:
                    dispatcher.utter_message(text=f"Có lỗi trong cơ sở dữ liệu: {str(e)}")
        else:
            dispatcher.utter_message(text="Vui lòng cung cấp đường dẫn đến các file Excel.")

        return []

class ActionSendEmail(Action):
    def name(self) -> str:
        return "action_send_email"

    def run(self, dispatcher, tracker, domain):
        question = tracker.get_slot('question')  # Lấy câu hỏi từ slot
        teacher_email = "duongdongto@gmail.com"
        sender_email = "duong0023@gmail.com"  # Email của bạn
        sender_password = "ibammoxasknmyttc"  # Mật khẩu email của bạn

        # Tạo nội dung email
        msg = MIMEText(f"Câu hỏi từ sinh viên: {question}")
        msg['Subject'] = 'Câu hỏi từ sinh viên'
        msg['From'] = sender_email
        msg['To'] = teacher_email

        # Gửi email
        try:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(sender_email, sender_password)
                server.send_message(msg)
            dispatcher.utter_message(text="Câu hỏi của bạn đã được chuyển tới giáo viên phụ trách.")
        except Exception as e:
            dispatcher.utter_message(text=f"Không thể gửi email: {str(e)}")
        return []
# def get_student_info_from_db(student_id):
#     try:
#         connection = sqlite3.connect('student.db')
#         cursor = connection.cursor()

#         cursor.execute('''
#             SELECT dot, malop, monhoc, hoten, mssv, vangphep, vangkhongphep, tongsotiet, phantramvang, ngayvang 
#             FROM STUDENT 
#             WHERE mssv = ?
#         ''', (student_id,))
        
#         student_data = cursor.fetchone()
#         connection.close()

#         if student_data:
#             return {
#                 'malop': student_data[0],
#                 'monhoc': student_data[1],
#                 'hoten': student_data[2],
#                 'mssv': student_data[3],
#                 'vangphep': student_data[4],
#                 'vangkhongphep': student_data[5],
#                 'phantramvang': student_data[6],
#                 'ngayvang': student_data[7],
#             }
#         else:
#             return None
#     except sqlite3.Error as e:
#         print(f"Database Error: {e}")
#         return None

# def parse_user_input(user_input):
#     student_id_match = re.search(r'\b\d{5,10}\b', user_input)
#     if student_id_match:
#         student_id = student_id_match.group(0)
#         return {'type': 'id', 'value': student_id}
    
#     if 'nghỉ' in user_input or 'vắng' in user_input:
#         if 'phép' in user_input:
#             return {'type': 'leave', 'leave_type': 'permission'}
#         elif 'không phép' in user_input or 'không lý do' in user_input:
#             return {'type': 'leave', 'leave_type': 'non-permission'}
    
#     return None

# def chatbot():
#     global student_name, student_id, class_name
#     while True:
#         user_input = input("Bạn: ").lower()
#         if user_input in ["thoát", "exit", "bye", "quit"]:
#             break
#         print("Chatbot: Tạm biệt!")
        
#         parsed_input = parse_user_input(user_input)
        
#         if parsed_input:
#             if parsed_input['type'] == 'id':
#                 student_id = parsed_input['value']
#                 student_info = get_student_info_from_db(student_id)
                
#                 if student_info:
#                     response = (f"Tên: {student_info['hoten']}, "
#                                 f"Lớp: {student_info['malop']}, "
#                                 f"Môn học: {student_info['monhoc']}, "
#                                 f"Số ngày nghỉ phép: {student_info['vangphep']}, "
#                                 f"Số ngày nghỉ không phép: {student_info['vangkhongphep']}, "
#                                 f"Phần trăm vắng: {student_info['phantramvang']}%, "
#                                 f"Ngày vắng: {student_info['ngayvang']}")
#                 else:
#                     response = f"Không tìm thấy thông tin sinh viên với ID {student_id}."
#                 print(f"Chatbot: {response}")
            
#             elif parsed_input['type'] == 'leave':
#                 leave_type = parsed_input['leave_type']
#                 if leave_type == 'permission':
#                     print("Chatbot: Bạn đang hỏi về ngày nghỉ phép. Vui lòng cung cấp thêm thông tin sinh viên hoặc ID.")
#                 elif leave_type == 'non-permission':
#                     print("Chatbot: Bạn đang hỏi về ngày nghỉ không phép. Vui lòng cung cấp thêm thông tin sinh viên hoặc ID.")
        
#         else:
#             # Nếu không hiểu câu hỏi, gửi email
#             send_email(["duongdongto@gmail.com"], "Câu hỏi từ sinh viên: ", user_input)
#             print("Chatbot: Xin lỗi, tôi chưa hiểu câu hỏi của bạn. Câu hỏi của bạn đã được gửi đến giáo viên.")

# def start_chatbot_thread():
#     chatbot_thread = threading.Thread(target=chatbot)
#     chatbot_thread.daemon = True
#     chatbot_thread.start()

# if __name__ == "__main__":
#     student_name = input("Nhập tên sinh viên: ")
#     student_id = input("Nhập mã sinh viên: ")
#     class_name = input("Nhập lớp: ")

#     start_chatbot_thread()

#     # while True:
#     #     command = input("Nhập lệnh chương trình chính (hoặc 'thoát' để kết thúc): ")
#     #     if command.lower() == 'thoát' or command.lower() == 'exit':
#     #         print("Chương trình chính kết thúc.")
#     #         break
#     #     else:
#     #         print(f"Chương trình chính đã nhận lệnh: {command}")

# Gán sự kiện cho Treeview
student_table.bind("<ButtonRelease-1>", select_item)
student_table.bind("<<TreeviewSelect>>", on_student_select)

load_data()

window.mainloop()
